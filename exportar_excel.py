# -*- coding: utf-8 -*-
"""Gera arquivos .xlsx a partir de colunas/linhas ja filtradas e
ordenadas pelo front-end, para o endpoint POST /api/exportar."""

import io
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

_RE_INVALIDO_NOME_ARQUIVO = re.compile(r"[^A-Za-z0-9_\-]+")
_RE_INVALIDO_TITULO_ABA = re.compile(r'[\[\]:*?/\\]')


def _nome_arquivo(titulo):
    limpo = _RE_INVALIDO_NOME_ARQUIVO.sub("_", titulo or "planilha").strip("_")
    hoje = datetime.now().strftime("%Y-%m-%d")
    return f"{limpo or 'planilha'}_{hoje}.xlsx"


def _valor_celula(bruto, tipo):
    if bruto is None:
        return None
    if tipo == "numero":
        try:
            return float(bruto)
        except (TypeError, ValueError):
            return None
    if tipo == "data":
        try:
            return datetime.fromisoformat(
                str(bruto).replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return str(bruto)
    return str(bruto)


def gerar(payload):
    """(bytes_do_xlsx, nome_do_arquivo) a partir de
    {"titulo": str, "colunas": [{"chave", "rotulo", "tipo"}], "linhas": [dict]}."""
    titulo = payload.get("titulo") or "Planilha"
    colunas = payload.get("colunas") or []
    linhas = payload.get("linhas") or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (_RE_INVALIDO_TITULO_ABA.sub("_", titulo) or "Planilha")[:31]

    for j, col in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=j, value=col.get("rotulo", ""))
        celula.font = Font(bold=True)

    for i, linha in enumerate(linhas, start=2):
        for j, col in enumerate(colunas, start=1):
            valor = _valor_celula(linha.get(col.get("chave")),
                                  col.get("tipo", "texto"))
            celula = ws.cell(row=i, column=j, value=valor)
            if col.get("tipo") == "data" and isinstance(valor, datetime):
                celula.number_format = "DD/MM/YYYY"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), _nome_arquivo(titulo)
