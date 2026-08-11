# -*- coding: utf-8 -*-
"""
Prototipo: processa os demonstrativos de pagamento recebidos por email
e monta um resumo por empresa e modalidade.

Fontes suportadas (pasta amostras/):
 - IDS: "Listagem de Repasse de Medico" (PDF) - totais por unidade/setor
 - Unimed Sorocaba: Demonstrativo de Pagamento (PDF) - servicos por codigo
 - CardioPro: planilha de repasse mensal (XLSX) - contagem ECG x MAPA

Uso:  py ler_repasses.py
"""

import glob
import json
import os
import re

import openpyxl
from pypdf import PdfReader

PASTA = "amostras"
ARQ_CONFIG_LOCAL = "config.json"


def pasta_documentos():
    """Pasta local extra definida pelo usuario (config.json), se valida."""
    try:
        with open(ARQ_CONFIG_LOCAL, "r", encoding="utf-8-sig") as f:
            p = (json.load(f).get("pasta_documentos") or "").strip()
        return p if p and os.path.isdir(p) else None
    except (OSError, ValueError):
        return None


ARQ_CACHE_PDF = "cache_pdf.json"
_cache_pdf = None


def _cache_pdf_carregado():
    global _cache_pdf
    if _cache_pdf is None:
        try:
            with open(ARQ_CACHE_PDF, "r", encoding="utf-8") as f:
                _cache_pdf = json.load(f)
        except (OSError, ValueError):
            _cache_pdf = {}
    return _cache_pdf


def texto_pdf(caminho):
    """Texto extraido do PDF, com cache em disco por caminho+tamanho+mtime.
    Os PDFs de repasse/exames sao grandes (as vezes 200+ paginas) e imutaveis
    depois de baixados, e o mesmo arquivo e lido varias vezes por carregamento
    (financeiro, cruzamento de pagamentos, etc.), entao vale a pena persistir."""
    cache = _cache_pdf_carregado()
    stat = os.stat(caminho)
    chave = os.path.abspath(caminho)
    assinatura = [stat.st_mtime, stat.st_size]
    entrada = cache.get(chave)
    if entrada and entrada.get("assinatura") == assinatura:
        return entrada["texto"]
    texto = "\n".join(p.extract_text() or "" for p in PdfReader(caminho).pages)
    cache[chave] = {"assinatura": assinatura, "texto": texto}
    with open(ARQ_CACHE_PDF, "w", encoding="utf-8") as f:
        json.dump(cache, f)
    return texto


def dinheiro(s):
    return float(s.replace(".", "").replace(",", "."))


# ---------------------------------------------------------------- IDS
def processar_ids(caminho):
    """Le totais por unidade/setor da Listagem de Repasse da IDS."""
    txt = texto_pdf(caminho)
    if "Listagem de Repasse" not in txt:
        return None
    resumo = {"tipo": "IDS - Listagem de Repasse", "arquivo": os.path.basename(caminho),
              "setores": [], "total": None}
    m = re.search(r"Listagem de Repasse de M.dico\s+Data: (\d{2}/\d{2}/\d{4})", txt)
    if m:
        resumo["emitido_em"] = m.group(1)
    unidade = None
    setor = None
    for linha in txt.splitlines():
        linha = linha.strip()
        # na extracao do PDF a ordem sai "Qtd Exames: N Total ...: R$ V"
        m = re.match(r"Unidade:\s*(.+)", linha)
        if m and "Todas" not in linha:
            unidade = m.group(1).strip()
            continue
        m = re.match(r"Setor:\s*(.+)", linha)
        if m:
            setor = m.group(1).strip()
            continue
        m = re.search(r"Qtd Exames:\s*(\d+)\s*Total para o setor:.*?"
                      r"R\$\s*([\d\.,]+)", linha)
        if m:
            resumo["setores"].append({
                "unidade": unidade, "setor": setor,
                "qtd": int(m.group(1)), "valor": dinheiro(m.group(2)),
            })
            continue
        m = re.search(r"Qtd Exames:\s*(\d+)\s*Total\s*R\$\s*([\d\.,]+)", linha)
        if m:
            resumo["total"] = {"qtd": int(m.group(1)),
                               "valor": dinheiro(m.group(2))}
    return resumo


# ---------------------------------------------------------------- Unimed
SERVICOS_UNIMED = {
    "20102038": "MAPA",
    "20102020": "Holter",
    "40101010": "ECG",
    "10101012": "Consulta",
    "99910073": "Consulta urgencia",
    "20101201": "Aval. marca-passo",
}


def processar_unimed(caminho):
    txt = texto_pdf(caminho)
    if "DEMONSTRATIVO" not in txt.upper() or "UNIMED" not in txt.upper():
        return None
    resumo = {"tipo": "Unimed - Demonstrativo", "arquivo": os.path.basename(caminho),
              "executantes": [], "servicos": {}, "liquido": None}
    m = re.search(r"Per.odo\s*(\d{6})", txt)
    if m:
        resumo["periodo"] = m.group(1)
    for m in re.finditer(
            r"Total prestador:\s*(\d+)\s+(.+?)\s+Servi.os:\s*(\d+)\s*R\$\s*([\d\.,]+)",
            txt):
        resumo["executantes"].append({
            "codigo": m.group(1), "nome": m.group(2).strip(),
            "servicos": int(m.group(3)), "valor": dinheiro(m.group(4)),
        })
    for codigo, nome in SERVICOS_UNIMED.items():
        n = len(re.findall(rf"\b{codigo}-", txt))
        if n:
            resumo["servicos"][nome] = n
    resumo["liquido"] = _valor_na_linha(txt, "total líquido")
    resumo["bruto"] = _valor_na_linha(txt, "emissão da nota")
    return resumo


def _valor_na_linha(txt, chave):
    """Valor R$ na linha que contem a chave, ou na linha seguinte."""
    linhas = txt.splitlines()
    for i, linha in enumerate(linhas):
        if chave.lower() not in linha.lower():
            continue
        proxima = linhas[i + 1] if i + 1 < len(linhas) else ""
        for cand in (linha, proxima):
            m = re.search(r"R\$\s*([\d\.,]+)", cand)
            if m:
                return dinheiro(m.group(1))
    return None


# ------------------------------------------------------- IDS Exames/Laudos
SETOR_EXAME_LISTAGEM = {
    "ECG": "ELETROCARDIOGRAMA IDS",
    "ELETROCARDIOGRAMA": "ELETROCARDIOGRAMA",
    "HONORARIO MEDICO": "LAUDO STRESS FARMACOLOGICO",
    "TESTE ERGOMETRICO": "TESTE ERGOMETRICO",
    "TESTE ERGOMETRICO MIBI": "TESTE ERGOMETRICO MIBI",
}

_FLAG = r"(?:N\s*ã\s*o|S\s*i\s*m)"
LINHA_LISTAGEM_RE = re.compile(
    r"^(?P<data>\d{2}/\d{2}/\d{4})\s+(?P<req>\d+)\s+(?P<pac>\d+)\s+(?P<meio>.+?)\s+"
    rf"(?P<f_ditado>{_FLAG})\s+(?P<f_digitado>{_FLAG})\s+(?P<f_impresso>{_FLAG})\s+"
    rf"(?P<f_assinado>{_FLAG})\s+(?P<f_liberado>{_FLAG})\s+(?P<f_entregue>{_FLAG})\s+"
    r"(?P<data_externa>\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})\s+"
    r"(?P<data_interna>\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})\s*$"
)


def _norm_flag(s):
    """Sim/Nao, tirando espacos que a extracao do PDF as vezes insere no meio."""
    return re.sub(r"\s+", "", s)


def _regex_tolerante(texto):
    """Regex que casa 'texto' mesmo com espacos extras inseridos no meio das palavras."""
    partes = []
    for c in texto:
        partes.append(r"\s+" if c == " " else re.escape(c) + r"\s*")
    return "".join(partes)


def _data_iso(data_br):
    d, m, a = data_br.split("/")
    return f"{a}-{m}-{d}"


def processar_listagem_exames(caminho):
    """Le a Listagem de Exames/Laudos da IDS: exames realizados com status do laudo
    (Ditado/Digitado/Impresso/Assinado/Liberado/Entregue), agrupados por setor."""
    txt = texto_pdf(caminho)
    if "Listagem de Exames" not in txt or "Laudos" not in txt:
        return None
    resumo = {"tipo": "IDS - Listagem de Exames/Laudos",
              "arquivo": os.path.basename(caminho), "setores": [], "exames": []}
    m = re.search(r"Per.odo de (\d{2}/\d{2}/\d{4}) at. (\d{2}/\d{2}/\d{4})", txt)
    if m:
        resumo["periodo"] = f"{m.group(1)} a {m.group(2)}"
    m = re.search(r"Data: (\d{2}/\d{2}/\d{4})", txt)
    if m:
        resumo["emitido_em"] = m.group(1)
    setor_atual = None
    contagem = {}
    for linha in txt.splitlines():
        linha = linha.strip()
        m = re.match(r"Setor: (?!Selecionado)(.+)", linha)
        if m:
            setor_atual = m.group(1).strip()
            contagem.setdefault(setor_atual, {"total": 0, "laudado": 0, "pendente": 0})
            continue
        m = LINHA_LISTAGEM_RE.match(linha)
        if not m or not setor_atual:
            continue
        assinado = _norm_flag(m.group("f_assinado"))
        nome = m.group("meio")
        exame_canonico = SETOR_EXAME_LISTAGEM.get(setor_atual)
        if exame_canonico:
            nome = re.sub(_regex_tolerante(exame_canonico) + r"$", "", nome).strip()
        resumo["exames"].append({
            "data": _data_iso(m.group("data")),
            "requisicao": m.group("req"),
            "paciente_codigo": m.group("pac"),
            "paciente": nome.title(),
            "setor": setor_atual,
            "ditado": _norm_flag(m.group("f_ditado")),
            "digitado": _norm_flag(m.group("f_digitado")),
            "impresso": _norm_flag(m.group("f_impresso")),
            "assinado": assinado,
            "liberado": _norm_flag(m.group("f_liberado")),
            "entregue": _norm_flag(m.group("f_entregue")),
        })
        c = contagem[setor_atual]
        c["total"] += 1
        c["laudado" if assinado == "Sim" else "pendente"] += 1
    resumo["setores"] = [{"setor": s, **v} for s, v in contagem.items()]
    m = re.search(r"Total Geral de Exames: (\d+)", txt)
    if m:
        resumo["total"] = int(m.group(1))
    return resumo


# ------------------------------------------ IDS Relatorio de Repasses
PROCEDIMENTOS_RELATORIO = (
    "TESTE ERGOMETRICO MIBI",
    "LAUDO STRESS FARMACOLOGICO",
    "TESTE ERGOMETRICO",
    "ELETROCARDIOGRAMA",
    "M.A.P.A",
    "MAPA",
)
_PROC_ALT = "|".join(re.escape(p) for p in PROCEDIMENTOS_RELATORIO)
LINHA_RELATORIO_RE = re.compile(
    r"^(?P<data>\d{2}/\d{2}/\d{4})\s+(?P<pac>.+?)\s+"
    rf"(?P<proc>{_PROC_ALT})\s+"
    r"(?P<valor>[\d\.]+,\d{2})\s+(?P<req>\d+)\s+(?P<conv>.+?)\s+\d+\s*$")


def processar_relatorio_repasses(caminho):
    """Le o Relatorio de Repasses Medicos (sistema da Medicina Nuclear/IDS):
    pagamento por paciente com exame nominal, valor, requisicao e convenio."""
    txt = texto_pdf(caminho)
    if not re.search(r"Relat.rio de Repasses M.dicos", txt):
        return None
    resumo = {"tipo": "IDS - Relatorio de Repasses",
              "arquivo": os.path.basename(caminho), "itens": [], "total": None}
    m = re.search(r"Per.odo de pesquisa entre\s*:\s*"
                  r"(\d{2}/\d{2}/\d{4}) e (\d{2}/\d{2}/\d{4})", txt)
    if m:
        resumo["periodo"] = f"{m.group(1)} a {m.group(2)}"
    m = re.search(r"Impresso em\s*:\s*(\d{1,2} \w{3} \d{4})", txt)
    if m:
        resumo["emitido_em"] = m.group(1)
    for linha in txt.splitlines():
        m = LINHA_RELATORIO_RE.match(linha.strip())
        if not m:
            continue
        resumo["itens"].append({
            "data": _data_iso(m.group("data")),
            "paciente": m.group("pac").strip(),
            "procedimento": m.group("proc"),
            "valor": dinheiro(m.group("valor")),
            "requisicao": m.group("req"),
            "convenio": m.group("conv").strip(),
        })
    m = re.search(r"Total procedimentos\s*R\$\s*([\d\.,]+)\s*(\d+)\s*Quantidade",
                  txt)
    if m:
        resumo["total"] = {"qtd": int(m.group(2)), "valor": dinheiro(m.group(1))}
    return resumo


# ---------------------------------------------------------------- CardioPro
def processar_cardiopro(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if not any("repasse" in aba.lower() for aba in wb.sheetnames):
        return None  # planilha de outro formato (ex.: contas medicas IDS)
    resumo = {"tipo": "CardioPro - Planilha de repasse",
              "arquivo": os.path.basename(caminho), "meses": []}
    for aba in wb.sheetnames:
        ws = wb[aba]
        ecg = mapa = 0
        for row in ws.iter_rows(values_only=True):
            for i, c in enumerate(row):
                v = str(c).strip() if c is not None else ""
                if v == "40101010" and i == 0:
                    ecg += 1
                elif v == "20102038":
                    mapa += 1
        resumo["meses"].append({"mes": aba, "ecg": ecg, "mapa": mapa})
    return resumo


def _fmt_valor(v):
    """Valor em reais no padrao brasileiro (ex.: 11.282,34)."""
    return f"{v:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


NOMES_AMIGAVEIS = {
    "IDS - Listagem de Repasse": "IDS · Repasse por unidade",
    "IDS - Listagem de Exames/Laudos": "IDS · Exames e laudos",
    "IDS - Relatorio de Repasses": "IDS · Relatório de repasses",
    "Unimed - Demonstrativo": "Unimed · Demonstrativo",
    "CardioPro - Planilha de repasse": "CardioPro · Planilha",
}


def _macro(r):
    """Resumo de uma linha pro card da aba Importacoes."""
    tipo = r["tipo"]
    if tipo == "IDS - Listagem de Repasse":
        if r.get("total"):
            return f"{r['total']['qtd']} exames · R$ {_fmt_valor(r['total']['valor'])}"
        return f"{len(r['setores'])} setor(es)"
    if tipo == "IDS - Listagem de Exames/Laudos":
        partes = [f"{r.get('total', '?')} exames"]
        if r.get("periodo"):
            partes.append(f"período {r['periodo']}")
        return " · ".join(partes)
    if tipo == "IDS - Relatorio de Repasses":
        if r.get("total"):
            return (f"{r['total']['qtd']} procedimentos · "
                    f"R$ {_fmt_valor(r['total']['valor'])}")
        return f"{len(r['itens'])} procedimento(s)"
    if tipo.startswith("Unimed"):
        partes = []
        if r.get("liquido"):
            partes.append(f"R$ {_fmt_valor(r['liquido'])} líquido")
        if r.get("periodo"):
            partes.append(f"período {r['periodo']}")
        return " · ".join(partes) if partes else f"{len(r['executantes'])} executante(s)"
    tot_ecg = sum(m["ecg"] for m in r["meses"])
    tot_mapa = sum(m["mapa"] for m in r["meses"])
    return f"{len(r['meses'])} mes(es) · {tot_ecg} ECG · {tot_mapa} MAPA"


def _tentar_parsers(caminho):
    """Roda o(s) parser(es) da extensao do arquivo.
    Devolve (resumo, erro): resumo e None se nenhum parser reconheceu o
    conteudo (ou a extensao nao tem parser); erro e a mensagem de excecao
    se algum parser quebrou no meio do caminho."""
    ext = os.path.splitext(caminho)[1].lower()
    try:
        if ext == ".pdf":
            r = None
            for parser in (processar_ids, processar_unimed, processar_listagem_exames, processar_relatorio_repasses):
                r = parser(caminho)
                if r:
                    break
            return r, None
        if ext == ".xlsx":
            return processar_cardiopro(caminho), None
        return None, None
    except Exception as e:
        return None, str(e)


def pastas_padrao():
    """repasses/ (email) + pasta local do usuario; amostras/ como reserva."""
    if glob.glob(os.path.join("repasses", "*")):
        base = ("repasses",)
    else:
        base = ("amostras",)
    extra = pasta_documentos()
    if extra:
        base = base + (extra,)
    return base


def _assinatura_conteudo(r):
    """Chave de dedup baseada no que o parser leu, nao no nome do arquivo.
    O mesmo demonstrativo pode chegar 2x por email (reenvio) com nomes
    diferentes (prefixo de data); sem isso o total da aba Financeiro conta
    o mesmo exame/valor 2x."""
    corpo = {k: v for k, v in r.items() if k != "arquivo"}
    return json.dumps(corpo, sort_keys=True, default=str)


def coletar(pastas=None):
    """Processa todos os demonstrativos das pastas (dedup por conteudo)."""
    if pastas is None:
        pastas = pastas_padrao()
    docs = []
    vistos = set()
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(glob.escape(pasta), "*"))):
            r, erro = _tentar_parsers(caminho)
            if erro:
                print(f"ERRO ao processar {caminho}: {erro}")
                continue
            if r:
                chave = _assinatura_conteudo(r)
                if chave in vistos:
                    continue
                vistos.add(chave)
                docs.append(r)
    return docs


def financeiro(pastas=None):
    """Estrutura consolidada por empresa para o painel."""
    empresas = {}
    for r in coletar(pastas):
        if r["tipo"] == "IDS - Listagem de Repasse":
            emp = empresas.setdefault("IDS", {"documentos": []})
            emp["documentos"].append({
                "arquivo": r["arquivo"],
                "emitido_em": r.get("emitido_em"),
                "linhas": [{"tipo": s["setor"].title(),
                            "detalhe": s["unidade"].title(),
                            "qtd": s["qtd"], "valor": s["valor"]}
                           for s in r["setores"]],
                "total": r.get("total"),
            })
        elif r["tipo"] == "IDS - Listagem de Exames/Laudos":
            pass  # coberto pela aba "Por fornecedor" (/api/realizados_fornecedor)
        elif r["tipo"].startswith("Unimed"):
            emp = empresas.setdefault("Unimed", {"documentos": []})
            emp["documentos"].append({
                "arquivo": r["arquivo"],
                "periodo": r.get("periodo"),
                "linhas": [{"tipo": t, "qtd": n, "valor": None}
                           for t, n in r["servicos"].items()],
                "executantes": r["executantes"],
                "bruto": r.get("bruto"),
                "liquido": r.get("liquido"),
            })
        elif r["tipo"] == "IDS - Relatorio de Repasses":
            emp = empresas.setdefault("IDS", {"documentos": []})
            por_proc = {}
            for it in r["itens"]:
                p = por_proc.setdefault(it["procedimento"], {"qtd": 0, "valor": 0})
                p["qtd"] += 1
                p["valor"] += it["valor"]
            emp["documentos"].append({
                "arquivo": r["arquivo"],
                "emitido_em": r.get("emitido_em"),
                "periodo": r.get("periodo"),
                "linhas": [{"tipo": proc.title(), "qtd": p["qtd"],
                            "valor": p["valor"]}
                           for proc, p in por_proc.items()],
                "total": r.get("total"),
            })
        else:
            emp = empresas.setdefault("CardioPro", {"documentos": []})
            tot_ecg = sum(m["ecg"] for m in r["meses"])
            tot_mapa = sum(m["mapa"] for m in r["meses"])
            emp["documentos"].append({
                "arquivo": r["arquivo"],
                "meses": r["meses"],
                "linhas": [{"tipo": "ECG", "qtd": tot_ecg, "valor": None},
                           {"tipo": "MAPA", "qtd": tot_mapa, "valor": None}],
            })
    return {"empresas": empresas}


def _inspecionar_arquivo(caminho):
    """Classifica um arquivo pro card da aba Importacoes: ok, nao_identificado
    ou erro."""
    nome = os.path.basename(caminho)
    ext = os.path.splitext(nome)[1].lower()
    if ext not in (".pdf", ".xlsx"):
        return {"arquivo": nome, "status": "nao_identificado",
                "motivo": f'Extensão "{ext or "(sem extensão)"}" ainda não tem parser'}
    r, erro = _tentar_parsers(caminho)
    if erro:
        return {"arquivo": nome, "status": "erro", "motivo": erro}
    if r is None:
        return {"arquivo": nome, "status": "nao_identificado",
                "motivo": "Nenhum parser conhecido reconheceu o conteúdo deste arquivo"}
    try:
        return {"arquivo": nome, "status": "ok", "tipo": r["tipo"],
                "tipo_amigavel": NOMES_AMIGAVEIS.get(r["tipo"], r["tipo"]),
                "resumo": _macro(r)}
    except Exception as e:
        return {"arquivo": nome, "status": "erro", "motivo": str(e)}


def inventario_pasta(pasta):
    """Status de cada arquivo de uma pasta pra aba Importacoes."""
    if not pasta or not os.path.isdir(pasta):
        return []
    return [_inspecionar_arquivo(caminho)
            for caminho in sorted(glob.glob(os.path.join(glob.escape(pasta), "*")))
            if os.path.isfile(caminho)]


def importacoes(pasta_email="repasses"):
    """Estrutura pra aba Importacoes: pasta local (prioritaria) + pasta do email."""
    local = pasta_documentos()
    return {
        "local": {"pasta": local or "",
                  "arquivos": inventario_pasta(local) if local else []},
        "email": {"pasta": pasta_email, "arquivos": inventario_pasta(pasta_email)},
    }


def main():
    achados = coletar()

    for r in achados:
        print("=" * 70)
        print(f"{r['tipo']}  [{r['arquivo']}]")
        if r["tipo"] == "IDS - Listagem de Repasse":
            soma_q = soma_v = 0
            for s in r["setores"]:
                print(f"  {s['unidade']:22s} {s['setor']:28s} "
                      f"{s['qtd']:4d} exames  R$ {s['valor']:10,.2f}")
                soma_q += s["qtd"]
                soma_v += s["valor"]
            if r["total"]:
                ok = ("OK" if (soma_q == r["total"]["qtd"]
                               and abs(soma_v - r["total"]["valor"]) < 0.01)
                      else "DIVERGENTE")
                print(f"  TOTAL DO DOCUMENTO: {r['total']['qtd']} exames "
                      f"R$ {r['total']['valor']:,.2f}  "
                      f"[conferencia soma dos setores: {ok}]")
        elif r["tipo"] == "IDS - Listagem de Exames/Laudos":
            print(f"  Periodo: {r.get('periodo', '?')}   "
                  f"Emitido: {r.get('emitido_em', '?')}")
            soma = 0
            for s in r["setores"]:
                print(f"  {s['setor']:24s} total {s['total']:4d}  "
                      f"laudados {s['laudado']:4d}  pendentes {s['pendente']:4d}")
                soma += s["total"]
            ok = "OK" if soma == r.get("total") else "DIVERGENTE"
            print(f"  TOTAL GERAL: {r.get('total')} exames  "
                  f"[conferencia soma dos setores: {ok}]")
        elif r["tipo"].startswith("Unimed"):
            print(f"  Periodo: {r.get('periodo', '?')}")
            for e in r["executantes"]:
                print(f"  Executante {e['nome']:30s} {e['servicos']:4d} "
                      f"servicos  R$ {e['valor']:10,.2f}")
            print(f"  Servicos por tipo: {r['servicos']}")
            if r.get("bruto"):
                print(f"  Bruto p/ nota: R$ {r['bruto']:,.2f}   "
                      f"Liquido: R$ {r['liquido']:,.2f}")
        else:
            for mes in r["meses"]:
                print(f"  {mes['mes']:28s} ECG: {mes['ecg']:4d}   "
                      f"MAPA: {mes['mapa']:4d}")


if __name__ == "__main__":
    main()
