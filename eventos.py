# -*- coding: utf-8 -*-
"""Modelo central de eventos de pagamento.

Todo demonstrativo (IDS, Unimed, CardioPro) e reduzido a uma lista unica de
eventos {pagador, exame, paciente, data, valor, convenio, tipo, documento}.
As agregacoes do painel (por mes, por pagador, por exame) e a investigacao
de "sem pagamento" trabalham so em cima dessa lista.
"""

import difflib
import glob
import os
import re
from datetime import datetime, timedelta

import openpyxl

import rotina_pendencias as rp
from ler_repasses import (_regex_tolerante, texto_pdf, dinheiro,
                           SETOR_EXAME_LISTAGEM,
                           processar_relatorio_repasses)

# ordem de prioridade quando o mesmo exame aparece pago em mais de uma fonte
PRIORIDADE_PAGADOR = {"IDS": 0, "Unimed": 1, "CardioPro": 2}

# nomes de setor (IDS) e de procedimento (Relatorio de Repasses) -> canonico
EXAME_CANONICO = {
    "MAPA": "MAPA",
    "M.A.P.A": "MAPA",
    "TESTE ERGOMETRICO": "Teste Ergométrico",
    "TESTE ERGOMETRICO MIBI": "Teste Ergométrico MIBI",
    "HONORARIO MEDICO": "Laudo Stress Farmacológico",
    "LAUDO STRESS FARMACOLOGICO": "Laudo Stress Farmacológico",
    "ECG": "Eletrocardiograma",
    "ELETROCARDIOGRAMA": "Eletrocardiograma",
}

# codigos de servico dos demonstrativos Unimed/CardioPro -> canonico
CODIGO_UNIMED_CANONICO = {
    "20102038": "MAPA",
    "20102020": "Holter",
    "40101010": "Eletrocardiograma",
    "10101012": "Consulta",
    "99910073": "Consulta",
    "20101201": "Aval. marca-passo",
}


def exame_canonico(nome):
    """Nome canonico do exame; desconhecido passa como veio (title case)."""
    chave = rp.normalizar(nome)
    return EXAME_CANONICO.get(chave, (nome or "").strip().title())


# convenios conhecidos grudados no fim do campo "nome" da Listagem de
# Repasse da IDS (sem separador); descobertos minerando os sufixos
CONVENIO_DISPLAY = {
    "INTERMEDICA SAUDE S.A": "Intermédica",
    "INTERMEDICA- MEDIPLAN": "Intermédica",
    "CARTAO IDS MATRIZ": "Cartão IDS",
    "HAPVIDA - SOROCABA": "Hapvida",
    "BRADESCO OPERADORA PLANOS S/A": "Bradesco Operadora",
    "BRADESCO SAUDE": "Bradesco Saúde",
    "CAIXA ECONOMICA FEDERAL": "Caixa Econômica Federal",
    "SUL AMERICA": "Sul América",
    "PORTO SEGURO": "Porto Seguro",
    "CEPOS- EMPRESA": "Cepos",
    "UNIMED": "Unimed",
    "AMIL": "Amil",
    "APAS": "Apas",
    "CASSI": "Cassi",
    "CABESP": "Cabesp",
    "MARINHA": "Marinha",
}
_PADROES_CONVENIO = sorted(CONVENIO_DISPLAY, key=len, reverse=True)
_PADROES_CONVENIO = [(c, re.compile(_regex_tolerante(c) + r"$"))
                     for c in _PADROES_CONVENIO]


def separar_convenio(nome):
    """(nome sem o sufixo de convenio, convenio de exibicao ou None)."""
    for canonico, padrao in _PADROES_CONVENIO:
        m = padrao.search(nome or "")
        if m:
            return (nome[:m.start()].strip(), CONVENIO_DISPLAY[canonico])
    return ((nome or "").strip(), None)


# -------------------------------------- extratores de itens por paciente
def data_de(v):
    if hasattr(v, "year"):
        return datetime(v.year, v.month, v.day)
    try:
        return datetime.strptime(str(v)[:10], "%d/%m/%Y")
    except ValueError:
        return None


def data_valida(d):
    """Descarta datas absurdas digitadas errado nos demonstrativos."""
    return (d is not None
            and datetime(2024, 1, 1) <= d <= datetime.now() + timedelta(days=45))


def casa_nome(nome_email, nome_pag):
    """nome_pag pode vir com o convenio grudado no fim (IDS)."""
    a = rp.normalizar(nome_email or "")
    b = rp.normalizar(nome_pag or "")
    if len(a.split()) < 2 or not b:
        return False
    if b.startswith(a) or a in b:
        return True
    return difflib.SequenceMatcher(None, a, b[:len(a) + 4]).ratio() >= 0.85


def itens_ids(caminho):
    txt = texto_pdf(caminho)
    if "Listagem de Repasse" not in txt:
        return []
    itens = []
    for linha in txt.splitlines():
        m = re.match(r"(\d{2}/\d{2}/\d{4})\s+(.+?)\s*M\.A\.P\.A", linha.strip())
        if m:
            mv = re.search(r"R\$\s*([\d\.]+,\d{2})", linha)
            valor = (float(mv.group(1).replace(".", "").replace(",", "."))
                     if mv else None)
            itens.append({"empresa": "IDS", "mod": "MAPA",
                          "data": data_de(m.group(1)),
                          "nome": m.group(2),  # nome + convenio grudados
                          "valor": valor,
                          "origem": os.path.basename(caminho)})
    return itens


def itens_ids_setores(caminho):
    """Itens de pagamento por paciente da Listagem de Repasse da IDS, para os
    setores que NAO sao MAPA (Teste Ergometrico, MIBI, ECG, Laudo Stress...).
    Cabecalho da tabela no PDF: Data Paciente Convenio Exame Pre-Labore Quant."""
    txt = texto_pdf(caminho)
    if "Listagem de Repasse" not in txt:
        return []
    itens = []
    setor_atual = None
    for linha in txt.splitlines():
        linha = linha.strip()
        m = re.match(r"Setor: (?!Selecionado)(.+)", linha)
        if m:
            setor_atual = m.group(1).strip()
            continue
        if not setor_atual or setor_atual == "MAPA":
            continue
        exame_canon = SETOR_EXAME_LISTAGEM.get(setor_atual)
        if not exame_canon:
            continue
        padrao = (r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+" +
                  _regex_tolerante(exame_canon) +
                  r"\s+R\s*\$\s*([\d\.]+\s*,\s*\d{2})\s+\d+\s*$")
        m = re.match(padrao, linha)
        if not m:
            continue
        nome = m.group(2).strip()
        itens.append({
            "empresa": "IDS", "mod": setor_atual,
            "data": data_de(m.group(1)), "nome": nome,
            "valor": dinheiro(re.sub(r"\s+", "", m.group(3))),
            "origem": os.path.basename(caminho),
            "convenio": separar_convenio(nome)[1],
        })
    return itens


def itens_unimed(caminho):
    """No PDF da Unimed cada campo sai numa linha:
    ... NOME / UF / plano(A|E) / data / CODIGO-SERVICO / qt / R$ ... / R$ pago"""
    txt = texto_pdf(caminho)
    if "UNIMED" not in txt.upper():
        return []
    linhas = [l.strip() for l in txt.splitlines()]
    itens = []
    for i, linha in enumerate(linhas):
        m = re.match(r"^(\d{8})-", linha)
        if not m or i < 4:
            continue
        codigo = m.group(1)
        mod = CODIGO_UNIMED_CANONICO.get(codigo)
        if not mod:
            continue
        data = data_de(linhas[i - 1])
        if not data:
            continue
        # nome: linhas alfabeticas logo acima do codigo de plano (i-3)
        partes = []
        j = i - 4
        while j >= 0 and len(partes) < 3:
            cand = linhas[j]
            if not cand or re.search(r"\d", cand):
                break  # protocolo/lote ou vazio: acabou o nome
            partes.insert(0, cand)
            j -= 1
        nome = " ".join(partes).strip()
        if len(nome.split()) < 2:
            continue
        # valor pago: ultimo "R$ ..." antes do proximo item
        valor = None
        for l2 in linhas[i + 1:i + 9]:
            if re.match(r"^\d{8}-", l2):
                break
            mv = re.match(r"^R\$\s*([\d\.,]+)$", l2)
            if mv:
                valor = dinheiro(mv.group(1))
        itens.append({"empresa": "Unimed", "mod": mod, "data": data,
                      "nome": nome, "valor": valor, "convenio": None,
                      "origem": os.path.basename(caminho)})
    return itens


def itens_cardiopro(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if not any("repasse" in aba.lower() for aba in wb.sheetnames):
        return []  # planilha de outro formato
    itens = []
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows(values_only=True):
            cells = list(row)
            for i, c in enumerate(cells):
                v = str(c).strip() if c is not None else ""
                if v != "20102038":
                    continue
                nome = cells[i + 1] if i + 1 < len(cells) else None
                nome = str(nome).strip() if nome else ""
                if len(nome.split()) < 2:
                    continue
                data = None
                for c2 in cells[i + 2:i + 6]:
                    d = data_de(c2) if c2 is not None else None
                    if d:
                        data = d
                        break
                itens.append({"empresa": "CardioPro", "mod": "MAPA",
                              "data": data, "nome": nome, "valor": None,
                              "convenio": None,
                              "origem": f"{os.path.basename(caminho)} [{aba}]"})
    return itens


def itens_relatorio(caminho):
    """Itens por paciente do Relatorio de Repasses Medicos (IDS)."""
    r = processar_relatorio_repasses(caminho)
    if not r:
        return []
    return [{"empresa": "IDS", "mod": it["procedimento"],
             "data": datetime.strptime(it["data"], "%Y-%m-%d"),
             "nome": it["paciente"], "valor": it["valor"],
             "convenio": separar_convenio(it["convenio"])[1] or it["convenio"].title(),
             "origem": r["arquivo"]}
            for it in r["itens"]]


# -------------------------------------------------- eventos consolidados
def _suprimir_cruzados(eventos):
    """O mesmo exame pago/faturado em fontes de pagadores diferentes conta
    uma vez, pela fonte de maior prioridade (IDS > Unimed > CardioPro).
    Casamento aproximado (casa_nome, +-10 dias) porque as fontes grafam o
    nome de formas diferentes.

    Tambem suprime reenvio/reimpressao do MESMO pagador quando o nome saiu
    truncado (ex.: "Filomena Marinho De Souz" x "Filomena Marinho De Souza
    Santos") ou grudado a um convenio desconhecido nao coberto por
    CONVENIO_DISPLAY (ex.: "Sandro Cravo Soares Votorantim Cimento Br -
    Empre" x "Sandro Cravo Soares"). Janela de data bem mais apertada
    (0-1 dia) que o cruzamento entre pagadores, pra nao juntar dois
    pacientes homonimos genuinamente distintos."""
    from datetime import datetime as _dt
    ordenados = sorted(eventos,
                       key=lambda e: PRIORIDADE_PAGADOR.get(e["pagador"], 9))
    mantidos = []
    indice = {}
    for evd in ordenados:
        tokens = rp.normalizar(evd["paciente"]).split()
        chave = (evd["exame"], tokens[0]) if tokens else None
        duplicado = False
        if chave and evd["data"]:
            dt = _dt.strptime(evd["data"], "%Y-%m-%d")
            for outro in indice.get(chave, []):
                if not outro["data"]:
                    continue
                mesmo_pagador = outro["pagador"] == evd["pagador"]
                delta = abs((_dt.strptime(outro["data"], "%Y-%m-%d") - dt).days)
                if delta > (1 if mesmo_pagador else 10):
                    continue
                if casa_nome(evd["paciente"], outro["paciente"]):
                    duplicado = True
                    break
        if duplicado:
            continue
        mantidos.append(evd)
        if chave:
            indice.setdefault(chave, []).append(evd)
    return mantidos


def coletar_eventos(pastas=None):
    """Lista unica de eventos de pagamento de todas as pastas, deduplicada."""
    import ler_repasses as lr
    if pastas is None:
        pastas = lr.pastas_padrao()
    brutos = []
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(glob.escape(pasta), "*"))):
            low = caminho.lower()
            try:
                if low.endswith(".pdf"):
                    brutos += itens_relatorio(caminho)
                    brutos += itens_ids(caminho)
                    brutos += itens_ids_setores(caminho)
                    brutos += itens_unimed(caminho)
                elif low.endswith(".xlsx"):
                    brutos += itens_cardiopro(caminho)
            except Exception:
                continue  # arquivo quebrado aparece na aba Importacoes
    eventos = []
    for p in brutos:
        data = p.get("data")
        if data is not None and not data_valida(data):
            data = None
        # so a IDS grudou convenio no campo nome sem separador; nas outras
        # fontes (Unimed/CardioPro) o nome e so o paciente, e tentar separar
        # convenio ali arrisca cortar sobrenome real (ex.: "Marinha")
        if p["empresa"] == "IDS":
            nome, conv = separar_convenio(p["nome"])
        else:
            nome, conv = p["nome"], None
        eventos.append({
            "pagador": p["empresa"],
            "exame": exame_canonico(p["mod"]),
            "paciente": nome.title(),
            "data": data.strftime("%Y-%m-%d") if data else None,
            "valor": p.get("valor"),
            "convenio": p.get("convenio") or conv,
            "tipo": "faturado" if p["empresa"] == "CardioPro" else "pago",
            "documento": p["origem"],
        })
    vistos = set()
    unicos = []
    for evd in eventos:
        chave = (evd["pagador"], evd["exame"],
                 rp.normalizar(evd["paciente"]), evd["data"])
        if chave in vistos:
            continue
        vistos.add(chave)
        unicos.append(evd)
    return _suprimir_cruzados(unicos)


# ------------------------------------------------- painel de recebimentos
def _exames_realizados(pastas=None):
    """Exames da Listagem de Exames/Laudos da IDS (dedup por requisicao):
    a evidencia de 'foi feito' pros exames que nao passam pelo email."""
    import ler_repasses as lr
    if pastas is None:
        pastas = lr.pastas_padrao()
    exames = []
    vistos_req = set()
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(glob.escape(pasta), "*"))):
            if not caminho.lower().endswith(".pdf"):
                continue
            try:
                r = lr.processar_listagem_exames(caminho)
            except Exception:
                continue
            for e in (r or {}).get("exames", []):
                if e["requisicao"] in vistos_req:
                    continue
                vistos_req.add(e["requisicao"])
                exames.append({"paciente": e["paciente"],
                               "setor": e["setor"], "data": e["data"],
                               "assinado": e["assinado"]})
    return exames


def recebimentos(pastas=None):
    """Estrutura completa pro GET /api/recebimentos."""
    from datetime import datetime as _dt
    import ler_repasses as lr
    evs = coletar_eventos(pastas)

    por_pagador = {}
    por_exame = {}
    por_mes = {}
    total_valor = 0.0
    exames_qtd = consultas_qtd = 0
    for evd in evs:
        v = evd["valor"] or 0
        total_valor += v
        if evd["exame"] == "Consulta":
            consultas_qtd += 1
        else:
            exames_qtd += 1
        p = por_pagador.setdefault(evd["pagador"], {"qtd": 0, "valor": 0})
        p["qtd"] += 1
        p["valor"] += v
        x = por_exame.setdefault(evd["exame"], {"qtd": 0, "valor": 0})
        x["qtd"] += 1
        x["valor"] += v
        if evd["data"]:
            m = por_mes.setdefault(evd["data"][:7],
                                   {"mes": evd["data"][:7],
                                    "por_pagador": {}, "por_exame": {}})
            mp = m["por_pagador"].setdefault(evd["pagador"],
                                             {"qtd": 0, "valor": 0})
            mp["qtd"] += 1
            mp["valor"] += v
            mx = m["por_exame"].setdefault(evd["exame"], {"qtd": 0, "valor": 0})
            mx["qtd"] += 1
            mx["valor"] += v

    # cobertura por exame: janela de datas de exame ja paga
    cobertura = {}
    for evd in evs:
        if not evd["data"]:
            continue
        c = cobertura.setdefault(evd["exame"], [evd["data"], evd["data"]])
        c[0] = min(c[0], evd["data"])
        c[1] = max(c[1], evd["data"])

    # sem pagamento: realizado laudado sem evento casado
    indice = {}
    for evd in evs:
        tokens = rp.normalizar(evd["paciente"]).split()
        if tokens and evd["data"]:
            indice.setdefault((evd["exame"], tokens[0]), []).append(evd)
    sem_pagamento = []
    for ex in _exames_realizados(pastas):
        if ex["assinado"] != "Sim":
            continue
        exame = exame_canonico(ex["setor"])
        tokens = rp.normalizar(ex["paciente"]).split()
        dt = _dt.strptime(ex["data"], "%Y-%m-%d")
        achou = False
        for evd in (indice.get((exame, tokens[0]), []) if tokens else []):
            delta = abs((_dt.strptime(evd["data"], "%Y-%m-%d") - dt).days)
            if delta <= 10 and casa_nome(ex["paciente"], evd["paciente"]):
                achou = True
                break
        if achou:
            continue
        c = cobertura.get(exame)
        # janela de cobertura por mes (nao por dia exato): um pagamento
        # registrado no mes cobre qualquer exame realizado no mesmo mes
        forca = ("forte" if c and c[0][:7] <= ex["data"][:7] <= c[1][:7]
                 else "fraca")
        sem_pagamento.append({"paciente": ex["paciente"], "exame": exame,
                              "data": ex["data"],
                              "fonte": "Listagem de Exames/Laudos (IDS)",
                              "forca": forca})
    sem_pagamento.sort(key=lambda c: (c["forca"] != "forte", c["data"]))

    lista_exames = sorted(
        ({"exame": nome, **tot} for nome, tot in por_exame.items()),
        key=lambda x: (-x["qtd"], -x["valor"]))
    return {
        "totais": {"valor": total_valor, "exames": exames_qtd,
                   "consultas": consultas_qtd, "por_pagador": por_pagador},
        "por_mes": sorted(por_mes.values(), key=lambda m: m["mes"]),
        "por_exame": lista_exames,
        "eventos": evs,
        "sem_pagamento": sem_pagamento,
        "cobertura": {k: {"inicio": v[0], "fim": v[1]}
                      for k, v in cobertura.items()},
        "documentos": lr.financeiro(pastas)["empresas"],
    }
