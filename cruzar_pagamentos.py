# -*- coding: utf-8 -*-
"""
Prototipo de cruzamento fase 2 x fase 3:
exames laudados (email) x exames pagos (demonstrativos em amostras/).

Pareamento por EMPRESA + NOME (aproximado) + DATA (+-10 dias), ja que os
demonstrativos nao trazem o codigo .dmw do exame.

Uso: py cruzar_pagamentos.py [--dias 95]
"""

import argparse
import difflib
import glob
import os
import re
from datetime import datetime, timedelta

import openpyxl

import rotina_pendencias as rp
from ler_repasses import (texto_pdf, dinheiro, SETOR_EXAME_LISTAGEM,
                           _regex_tolerante, processar_listagem_exames,
                           pastas_padrao)
from eventos import separar_convenio

PASTA = "amostras"

# Regras da Dra. Gisele (26/07/2026):
# - exames de fonte IDS sao pagos pela IDS (raros casos via Unimed)
# - exames da Unimed (ecocardiograma) sao pagos pela Unimed
# - exames da CardioPro sao PAGOS PELA UNIMED; a planilha da CardioPro
#   e producao/faturamento, nao pagamento
# - TE/MIBI/stress farmacologico: pagamento exclusivo da IDS
COMPAT = {  # pagador -> fontes de exame que ele pode pagar
    "IDS": ("IDS",),
    "Unimed": ("Unimed", "CardioPro", "IDS"),
    "CardioPro": ("CardioPro",),
}
ORDEM_PAGADOR = {"IDS": 0, "Unimed": 1, "CardioPro": 2}
# janela de cobertura usada no alerta "sem registro", por fonte do exame
PAGADOR_PRIMARIO = {"IDS": "IDS", "Unimed": "Unimed", "CardioPro": "Unimed"}

def data_de(v):
    if hasattr(v, "year"):
        return datetime(v.year, v.month, v.day)
    try:
        return datetime.strptime(str(v)[:10], "%d/%m/%Y")
    except ValueError:
        return None


def data_valida(d):
    """Descarta datas absurdas digitadas errado nos demonstrativos."""
    return (d is not None
            and datetime(2024, 1, 1) <= d <= datetime.now() + timedelta(days=45))


# ---------------- pagamentos: extracao de itens (nome + data) ----------------
def itens_ids(caminho):
    txt = texto_pdf(caminho)
    if "Listagem de Repasse" not in txt:
        return []
    itens = []
    for linha in txt.splitlines():
        m = re.match(r"(\d{2}/\d{2}/\d{4})\s+(.+?)\s*M\.A\.P\.A", linha.strip())
        if m:
            mv = re.search(r"R\$\s*([\d\.]+,\d{2})", linha)
            valor = (float(mv.group(1).replace(".", "").replace(",", "."))
                     if mv else None)
            itens.append({"empresa": "IDS", "mod": "MAPA",
                          "data": data_de(m.group(1)),
                          "nome": m.group(2),  # nome + convenio grudados
                          "valor": valor,
                          "origem": os.path.basename(caminho)})
    return itens


def itens_ids_setores(caminho):
    """Itens de pagamento por paciente da Listagem de Repasse da IDS, para os
    setores que NAO sao MAPA (Teste Ergometrico, MIBI, ECG, Laudo Stress...).
    Cabecalho da tabela no PDF: Data Paciente Convenio Exame Pre-Labore Quant."""
    txt = texto_pdf(caminho)
    if "Listagem de Repasse" not in txt:
        return []
    itens = []
    setor_atual = None
    for linha in txt.splitlines():
        linha = linha.strip()
        m = re.match(r"Setor: (?!Selecionado)(.+)", linha)
        if m:
            setor_atual = m.group(1).strip()
            continue
        if not setor_atual or setor_atual == "MAPA":
            continue
        exame_canonico = SETOR_EXAME_LISTAGEM.get(setor_atual)
        if not exame_canonico:
            continue
        padrao = (r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+" +
                  _regex_tolerante(exame_canonico) +
                  r"\s+R\s*\$\s*([\d\.]+\s*,\s*\d{2})\s+\d+\s*$")
        m = re.match(padrao, linha)
        if not m:
            continue
        nome = m.group(2).strip()
        itens.append({
            "empresa": "IDS", "mod": setor_atual,
            "data": data_de(m.group(1)), "nome": nome,
            "valor": dinheiro(re.sub(r"\s+", "", m.group(3))),
            "origem": os.path.basename(caminho),
            "convenio": separar_convenio(nome)[1],
        })
    return itens


def coletar_itens_setores(pastas=None):
    """Pagamentos por paciente dos setores nao-MAPA, deduplicados por
    setor+nome+data (independente do coletar_itens() usado pelo MAPA)."""
    if pastas is None:
        pastas = pastas_padrao()
    itens = []
    vistos_arq = set()
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(pasta, "*"))):
            if not caminho.lower().endswith(".pdf"):
                continue
            arq = os.path.basename(caminho).lower()
            if arq in vistos_arq:
                continue
            novos = itens_ids_setores(caminho)
            if novos:
                vistos_arq.add(arq)
                itens.extend(novos)
    vistos = set()
    unicos = []
    for p in itens:
        if p["data"] is not None and not data_valida(p["data"]):
            p["data"] = None
        chave = (p["mod"], rp.normalizar(p["nome"]), p["data"])
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(p)
    return unicos


def cruzar_realizados_ids(pastas=None):
    """Casa os exames da Listagem de Exames/Laudos (pedido + status do laudo)
    com os pagamentos por paciente da Listagem de Repasse, por setor.
    Devolve a lista de exames (dedup por requisicao) com 'pagamento' anotado
    quando encontrado, ou 'pagamento_esperado' quando laudado, dentro da
    janela de datas ja coberta pelos repasses daquele setor, e sem par."""
    if pastas is None:
        pastas = pastas_padrao()
    exames = []
    vistos_arq = set()
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(pasta, "*"))):
            if not caminho.lower().endswith(".pdf"):
                continue
            arq = os.path.basename(caminho).lower()
            if arq in vistos_arq:
                continue
            r = processar_listagem_exames(caminho)
            if r:
                vistos_arq.add(arq)
                exames.extend(r["exames"])

    vistos_req = set()
    unicos = []
    for e in exames:
        if e["requisicao"] in vistos_req:
            continue
        vistos_req.add(e["requisicao"])
        unicos.append(e)
    exames = unicos

    pagamentos = coletar_itens_setores(pastas)
    indice = {}
    for p in pagamentos:
        tokens = rp.normalizar(p["nome"]).split()
        if tokens:
            indice.setdefault((p["mod"], tokens[0]), []).append(p)

    cobertura = {}
    for p in pagamentos:
        if p["data"]:
            c = cobertura.setdefault(p["mod"], [p["data"], p["data"]])
            if p["data"] < c[0]:
                c[0] = p["data"]
            if p["data"] > c[1]:
                c[1] = p["data"]

    usados = set()
    for ex in exames:
        ex_data = datetime.strptime(ex["data"], "%Y-%m-%d")
        tokens = rp.normalizar(ex["paciente"]).split()
        achou = None
        for p in (indice.get((ex["setor"], tokens[0]), []) if tokens else []):
            if id(p) in usados or not p["data"]:
                continue
            if abs((p["data"] - ex_data).days) > 10:
                continue
            if casa_nome(ex["paciente"], p["nome"]):
                achou = p
                break
        if achou:
            usados.add(id(achou))
            ex["pagamento"] = {"data": achou["data"].strftime("%Y-%m-%d"),
                               "valor": achou.get("valor"),
                               "origem": achou["origem"],
                               "convenio": achou.get("convenio")}
        else:
            c = cobertura.get(ex["setor"])
            if ex["assinado"] == "Sim" and c and c[0] <= ex_data <= c[1]:
                ex["pagamento_esperado"] = True
    return exames


def agregar_por_mes_fornecedor(pastas=None):
    """Agrega os exames (realizados/laudados/pagos) por mes e, na parte paga,
    por convenio -- para o painel mostrar producao x pagamento por fornecedor.
    'Realizados'/'laudados' nao tem convenio (nao vem na Listagem de Exames/
    Laudos); so a fatia paga sabe de quem e, porque so descobrimos o convenio
    quando o pagamento casa."""
    exames = cruzar_realizados_ids(pastas)
    meses = {}
    fornecedores_total = {}
    for ex in exames:
        mes = ex["data"][:7]
        m = meses.setdefault(mes, {"mes": mes, "realizados": 0, "laudados": 0,
                                    "pago_total": 0, "por_fornecedor": {},
                                    "nao_laudados": [], "laudados_sem_pagamento": []})
        m["realizados"] += 1
        if ex["assinado"] == "Sim":
            m["laudados"] += 1
        else:
            m["nao_laudados"].append({"paciente": ex["paciente"], "data": ex["data"],
                                       "setor": ex["setor"]})
        pag = ex.get("pagamento")
        if pag:
            m["pago_total"] += 1
            conv = pag.get("convenio") or "Outros"
            fc = m["por_fornecedor"].setdefault(
                conv, {"qtd": 0, "valor": 0, "transacoes": []})
            fc["qtd"] += 1
            fc["valor"] += pag.get("valor") or 0
            fc["transacoes"].append({
                "paciente": ex["paciente"],
                "data": ex["data"],
                "setor": ex["setor"],
                "valor": pag.get("valor"),
                "data_pagamento": pag.get("data"),
                "origem": pag.get("origem"),
            })
            ft = fornecedores_total.setdefault(conv, {"qtd": 0, "valor": 0})
            ft["qtd"] += 1
            ft["valor"] += pag.get("valor") or 0
        elif ex["assinado"] == "Sim":
            m["laudados_sem_pagamento"].append({
                "paciente": ex["paciente"], "data": ex["data"], "setor": ex["setor"],
                "esperado": bool(ex.get("pagamento_esperado")),
            })

    total_pago_geral = sum(f["qtd"] for f in fornecedores_total.values())
    fornecedores = [
        {"nome": nome, "qtd": f["qtd"], "valor": f["valor"],
         "participacao": f["qtd"] / total_pago_geral if total_pago_geral else 0}
        for nome, f in fornecedores_total.items()
    ]
    fornecedores.sort(key=lambda f: -f["qtd"])
    return {"meses": sorted(meses.values(), key=lambda m: m["mes"]),
            "fornecedores": fornecedores}


def itens_unimed(caminho):
    """No PDF da Unimed cada campo sai numa linha:
    ... NOME / UF / plano(A|E) / data / 20102038-SERVICO"""
    txt = texto_pdf(caminho)
    if "UNIMED" not in txt.upper():
        return []
    linhas = [l.strip() for l in txt.splitlines()]
    itens = []
    for i, linha in enumerate(linhas):
        if not linha.startswith("20102038") or i < 4:
            continue
        data = data_de(linhas[i - 1])
        if not data:
            continue
        # nome: linhas alfabeticas logo acima do codigo de plano (i-3)
        partes = []
        j = i - 4
        while j >= 0 and len(partes) < 3:
            cand = linhas[j]
            if not cand or re.search(r"\d", cand):
                break  # protocolo/lote ou vazio: acabou o nome
            partes.insert(0, cand)
            j -= 1
        nome = " ".join(partes).strip()
        if len(nome.split()) >= 2:
            itens.append({"empresa": "Unimed", "mod": "MAPA", "data": data,
                          "nome": nome,
                          "origem": os.path.basename(caminho)})
    return itens


def itens_cardiopro(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if not any("repasse" in aba.lower() for aba in wb.sheetnames):
        return []  # planilha de outro formato
    itens = []
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows(values_only=True):
            cells = list(row)
            for i, c in enumerate(cells):
                v = str(c).strip() if c is not None else ""
                if v != "20102038":
                    continue
                nome = cells[i + 1] if i + 1 < len(cells) else None
                nome = str(nome).strip() if nome else ""
                if len(nome.split()) < 2:
                    continue
                data = None
                for c2 in cells[i + 2:i + 6]:
                    d = data_de(c2) if c2 is not None else None
                    if d:
                        data = d
                        break
                itens.append({"empresa": "CardioPro", "mod": "MAPA",
                              "data": data, "nome": nome,
                              "origem": f"{os.path.basename(caminho)} [{aba}]"})
    return itens


# ---------------- pareamento ----------------
def casa_nome(nome_email, nome_pag):
    """nome_pag pode vir com o convenio grudado no fim (IDS)."""
    a = rp.normalizar(nome_email or "")
    b = rp.normalizar(nome_pag or "")
    if len(a.split()) < 2 or not b:
        return False
    if b.startswith(a) or a in b:
        return True
    return difflib.SequenceMatcher(None, a, b[:len(a) + 4]).ratio() >= 0.85


def coletar_itens(pastas=None):
    """Todos os itens de pagamento MAPA dos demonstrativos, deduplicados."""
    from ler_repasses import pastas_padrao
    if pastas is None:
        pastas = pastas_padrao()
    itens = []
    vistos_arq = set()
    for pasta in pastas:
        for caminho in sorted(glob.glob(os.path.join(pasta, "*"))):
            arq = os.path.basename(caminho).lower()
            if arq in vistos_arq:
                continue
            if caminho.lower().endswith(".pdf"):
                novos = itens_ids(caminho) or itens_unimed(caminho)
            elif caminho.lower().endswith(".xlsx"):
                novos = itens_cardiopro(caminho)
            else:
                continue
            if novos:
                vistos_arq.add(arq)
                itens.extend(novos)
    vistos = set()
    unicos = []
    for p in itens:
        if p["data"] is not None and not data_valida(p["data"]):
            p["data"] = None  # data digitada errado no demonstrativo
        chave = (p["empresa"], rp.normalizar(p["nome"]), p["data"])
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(p)
    return unicos


def anotar_pagamentos(dados):
    """Anota ex['pagamento'] nos exames de `dados` (retorno de analisar) e
    devolve a lista de pagamentos sem exame correspondente na janela."""
    itens = coletar_itens()
    exames = []
    for lista in ("retornados", "pendentes", "provaveis", "avisos", "baixados"):
        exames.extend(dados[lista])

    indice = {}
    for ex in exames:
        if not ex.get("nome"):
            continue
        ex["_dt"] = datetime.fromisoformat(
            ex["recebido"].replace("Z", "+00:00")).replace(tzinfo=None)
        chave = (ex["empresa"], rp.normalizar(ex["nome"]).split()[0])
        indice.setdefault(chave, []).append(ex)

    # pagamentos reais (IDS, Unimed) tem prioridade sobre a planilha de
    # producao da CardioPro na hora de reivindicar um exame
    itens.sort(key=lambda p: ORDEM_PAGADOR.get(p["empresa"], 9))

    # janela de cobertura por empresa: intervalo de datas de exame que os
    # demonstrativos ja pagaram; exame dentro dela sem pagamento = em falta
    cobertura = {}
    for p in itens:
        if p["data"]:
            c = cobertura.setdefault(p["empresa"], [p["data"], p["data"]])
            if p["data"] < c[0]:
                c[0] = p["data"]
            if p["data"] > c[1]:
                c[1] = p["data"]
    dados["cobertura_pagamentos"] = {
        emp: {"inicio": c[0].strftime("%Y-%m-%d"),
              "fim": c[1].strftime("%Y-%m-%d")}
        for emp, c in cobertura.items()
    }

    inicio = min((ex["_dt"] for ex in exames if "_dt" in ex), default=None)
    usados = set()
    orfaos = []
    for p in itens:
        if not p["data"] or not p["nome"]:
            continue
        tokens = rp.normalizar(p["nome"]).split()
        achou = None
        for fonte in COMPAT.get(p["empresa"], (p["empresa"],)):
            for ex in indice.get((fonte, tokens[0]), []) if tokens else []:
                if ex["codigo"] in usados:
                    continue
                if abs((ex["_dt"] - p["data"]).days) > 10:
                    continue
                if casa_nome(ex["nome"], p["nome"]):
                    achou = ex
                    break
            if achou:
                break
        if achou:
            usados.add(achou["codigo"])
            achou["pagamento"] = {
                "data": p["data"].strftime("%Y-%m-%d"),
                "valor": p.get("valor"),
                "origem": p["origem"],
                "pagador": p["empresa"],
                # planilha CardioPro comprova faturamento, nao pagamento
                "tipo": ("faturado" if p["empresa"] == "CardioPro"
                         else "pago"),
            }
        elif inicio and p["data"] >= inicio - timedelta(days=3):
            orfaos.append({"empresa": p["empresa"], "nome": p["nome"],
                           "data": p["data"].strftime("%Y-%m-%d"),
                           "origem": p["origem"]})
    for ex in exames:
        # so faz sentido cobrar pagamento de exame ja laudado; a janela
        # usada e a do pagador responsavel pela fonte do exame
        if ("_dt" in ex and not ex.get("pagamento")
                and ex.get("retornado_em")):
            pagador = PAGADOR_PRIMARIO.get(ex["empresa"])
            c = cobertura.get(pagador) if pagador else None
            if c and c[0] <= ex["_dt"] <= c[1]:
                ex["pagamento_esperado"] = True

    # regra da repeticao: mesmo paciente, mesma fonte, exames com <= 15
    # dias de intervalo e so um pagamento = o outro e repeticao (exame
    # que nao deu certo), nao pendencia
    por_paciente = {}
    for ex in exames:
        if ex.get("nome") and "_dt" in ex:
            chave = (ex["empresa"], rp.normalizar(ex["nome"]))
            por_paciente.setdefault(chave, []).append(ex)
    for grupo in por_paciente.values():
        if len(grupo) < 2:
            continue
        for ex in grupo:
            if not ex.get("pagamento_esperado"):
                continue
            for irmao in grupo:
                if (irmao is not ex and irmao.get("pagamento")
                        and abs((ex["_dt"] - irmao["_dt"]).days) <= 15):
                    ex["repeticao"] = True
                    ex.pop("pagamento_esperado", None)
                    break

    for ex in exames:
        ex.pop("_dt", None)
    orfaos.sort(key=lambda o: (o["empresa"], o["data"]))
    return orfaos


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dias", type=int, default=95)
    args = ap.parse_args()

    print(f"Lendo {args.dias} dias de emails (pode levar alguns minutos)...")
    dados = rp.analisar(args.dias)
    laudados = []
    for lista in ("retornados", "pendentes", "provaveis", "avisos", "baixados"):
        for ex in dados[lista]:
            ex["status"] = lista
            ex["recebido_dt"] = datetime.fromisoformat(
                ex["recebido"].replace("Z", "+00:00")).replace(tzinfo=None)
            laudados.append(ex)
    print(f"Exames no email ({args.dias}d): {len(laudados)}")

    pagos = []
    for caminho in sorted(glob.glob(os.path.join(PASTA, "*"))):
        if caminho.lower().endswith(".pdf"):
            pagos += itens_ids(caminho) or itens_unimed(caminho)
        elif caminho.lower().endswith(".xlsx"):
            pagos += itens_cardiopro(caminho)
    # demonstrativos repetem linhas do mesmo exame (CardioPro e Unimed
    # listam o MAPA 2x); deduplica por empresa + nome + data
    vistos = set()
    unicos = []
    for p in pagos:
        chave = (p["empresa"], rp.normalizar(p["nome"]), p["data"])
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(p)
    pagos = unicos
    print(f"Itens MAPA nos demonstrativos (apos dedup): {len(pagos)}")

    inicio_email = min(ex["recebido_dt"] for ex in laudados)
    janela = [p for p in pagos
              if p["data"] and p["data"] >= inicio_email - timedelta(days=3)]
    print(f"Itens pagos dentro da janela coberta pelos emails: {len(janela)}\n")

    por_empresa = {}
    for p in janela:
        por_empresa.setdefault(p["empresa"], []).append(p)

    for empresa, itens in sorted(por_empresa.items()):
        candidatos = [ex for ex in laudados if ex["empresa"] == empresa]
        casados, orfaos = [], []
        usados = set()
        for p in itens:
            achou = None
            for ex in candidatos:
                if ex["codigo"] in usados or not ex["nome"]:
                    continue
                if abs((ex["recebido_dt"] - p["data"]).days) > 10:
                    continue
                if casa_nome(ex["nome"], p["nome"]):
                    achou = ex
                    break
            if achou:
                usados.add(achou["codigo"])
                casados.append((p, achou))
            else:
                orfaos.append(p)
        print("=" * 70)
        print(f"{empresa}: {len(itens)} pagamentos na janela | "
              f"{len(casados)} casados com exame do email "
              f"({100 * len(casados) // max(1, len(itens))}%) | "
              f"{len(orfaos)} sem par")
        for p, ex in casados[:5]:
            print(f"  OK  {p['data']:%d/%m} {p['nome'][:38]:38s} -> "
                  f"{ex['codigo']} {ex['nome'][:30]} ({ex['status']})")
        if len(casados) > 5:
            print(f"  ... +{len(casados) - 5} casamentos")
        for p in orfaos[:8]:
            print(f"  ??  {p['data']:%d/%m} {p['nome'][:50]}  [{p['origem'][:30]}]")
        if len(orfaos) > 8:
            print(f"  ... +{len(orfaos) - 8} sem par")


if __name__ == "__main__":
    main()
