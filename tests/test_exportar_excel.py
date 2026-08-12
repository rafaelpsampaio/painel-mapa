import io
from datetime import datetime

import openpyxl

import exportar_excel as ee


def test_gerar_cria_cabecalho_e_linhas():
    payload = {
        "titulo": "Exames",
        "colunas": [
            {"chave": "codigo", "rotulo": "Código", "tipo": "texto"},
            {"chave": "qtd", "rotulo": "Qtd", "tipo": "numero"},
        ],
        "linhas": [
            {"codigo": "ED9-00159", "qtd": 3},
            {"codigo": "0RC-00042", "qtd": 1},
        ],
    }
    conteudo, nome_arquivo = ee.gerar(payload)
    assert nome_arquivo.startswith("Exames_")
    assert nome_arquivo.endswith(".xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Código", "Qtd"]
    assert [c.value for c in ws[2]] == ["ED9-00159", 3]
    assert [c.value for c in ws[3]] == ["0RC-00042", 1]


def test_gerar_coluna_numero_vira_float_de_verdade():
    payload = {
        "titulo": "Por exame",
        "colunas": [{"chave": "valor", "rotulo": "Valor", "tipo": "numero"}],
        "linhas": [{"valor": "123.45"}],
    }
    conteudo, _ = ee.gerar(payload)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    valor = wb.active.cell(row=2, column=1).value
    assert valor == 123.45
    assert isinstance(valor, float)


def test_gerar_coluna_numero_invalido_preserva_valor_como_texto():
    payload = {
        "titulo": "Por exame",
        "colunas": [{"chave": "valor", "rotulo": "Valor", "tipo": "numero"}],
        "linhas": [{"valor": "nao é um número"}],
    }
    conteudo, _ = ee.gerar(payload)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    valor = wb.active.cell(row=2, column=1).value
    assert valor == "nao é um número"


def test_gerar_coluna_data_vira_data_de_verdade():
    payload = {
        "titulo": "Eventos",
        "colunas": [{"chave": "recebido", "rotulo": "Recebido", "tipo": "data"}],
        "linhas": [{"recebido": "2026-08-01T10:00:00Z"}],
    }
    conteudo, _ = ee.gerar(payload)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    celula = wb.active.cell(row=2, column=1)
    assert celula.value == datetime(2026, 8, 1, 10, 0, 0)
    assert celula.number_format == "DD/MM/YYYY"


def test_gerar_valor_ausente_vira_celula_vazia():
    payload = {
        "titulo": "Exames",
        "colunas": [{"chave": "prazo", "rotulo": "Prazo", "tipo": "data"}],
        "linhas": [{"prazo": None}],
    }
    conteudo, _ = ee.gerar(payload)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert wb.active.cell(row=2, column=1).value is None


def test_nome_arquivo_sanitiza_titulo_com_espacos():
    payload = {"titulo": "Por exame", "colunas": [], "linhas": []}
    _, nome_arquivo = ee.gerar(payload)
    assert nome_arquivo.startswith("Por_exame_")
    assert " " not in nome_arquivo
